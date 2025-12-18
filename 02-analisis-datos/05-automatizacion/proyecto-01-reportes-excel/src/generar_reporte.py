"""
Generador Automático de Reportes en Excel
==========================================

Automatiza la creación de reportes ejecutivos en Excel con:
- Múltiples hojas
- Formato profesional
- Gráficos integrados
- Tablas dinámicas
- Estilos condicionales
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime

class ExcelReportGenerator:
    """Generador de reportes Excel automatizados."""
    
    def __init__(self):
        """Inicializa el generador."""
        self.# Buscar carpeta base \'data engineer\'
current_path = Path(__file__).resolve()
base_path = None
for parent in current_path.parents:
    if parent.name == \'data engineer\':
        base_path = parent
        break
if base_path is None:
    raise FileNotFoundError("No se pudo encontrar la carpeta base \'data engineer\'")
        self.data_path = self.base_path / '02-analisis-datos/01-eda-exploratorio/proyecto-01-ventas-retail/data/raw'
        self.output_path = Path(__file__).parent.parent / 'output'
        self.output_path.mkdir(exist_ok=True)
        
        # Cargar datos
        self.df = pd.read_csv(self.data_path / 'transacciones.csv')
        self.df['fecha'] = pd.to_datetime(self.df['fecha'])
        
        print(f"✅ Datos cargados: {len(self.df):,} transacciones")
    
    def crear_reporte_ejecutivo(self):
        """Crea reporte ejecutivo completo."""
        print("\n📊 Generando reporte ejecutivo...")
        
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remover sheet por defecto
        
        # 1. Hoja de Resumen
        self._crear_hoja_resumen(wb)
        
        # 2. Hoja de Ventas por Categoría
        self._crear_hoja_categorias(wb)
        
        # 3. Hoja de Tendencias
        self._crear_hoja_tendencias(wb)
        
        # 4. Hoja de Top Productos
        self._crear_hoja_top_productos(wb)
        
        # 5. Hoja de Datos Crudos (muestra)
        self._crear_hoja_datos(wb)
        
        # Guardar
        filename = f"Reporte_Ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_path / filename
        wb.save(filepath)
        
        print(f"   ✅ Reporte guardado: {filename}")
        return filepath
    
    def _crear_hoja_resumen(self, wb):
        """Crea hoja de resumen con KPIs."""
        ws = wb.create_sheet("📊 Resumen Ejecutivo")
        
        # Título
        ws['B2'] = "REPORTE EJECUTIVO DE VENTAS"
        ws['B2'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['B2'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws['B2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('B2:E2')
        
        # Fecha del reporte
        ws['B3'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B3'].font = Font(italic=True)
        
        # KPIs
        kpis = [
            ("Total Ingresos", f"€{self.df['total'].sum():,.2f}"),
            ("Total Transacciones", f"{len(self.df):,}"),
            ("Ticket Promedio", f"€{self.df['total'].mean():.2f}"),
            ("Productos Vendidos", f"{self.df['cantidad'].sum():,}")
        ]
        
        row = 5
        for i, (nombre, valor) in enumerate(kpis):
            # Nombre del KPI
            cell_nombre = ws.cell(row=row, column=2, value=nombre)
            cell_nombre.font = Font(bold=True, size=12)
            cell_nombre.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Valor del KPI
            cell_valor = ws.cell(row=row, column=3, value=valor)
            cell_valor.font = Font(size=14, bold=True, color="0070C0")
            cell_valor.alignment = Alignment(horizontal="right")
            
            row += 2
        
        # Ajustar anchos
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        
        print("   ✅ Hoja Resumen creada")
    
    def _crear_hoja_categorias(self, wb):
        """Crea hoja con análisis por categoría."""
        ws = wb.create_sheet("📦 Por Categoría")
        
        # Preparar datos
        df_cat = self.df.groupby('categoria').agg({
            'total': ['sum', 'mean', 'count'],
            'cantidad': 'sum'
        }).round(2)
        
        df_cat.columns = ['Ingresos Totales', 'Ticket Promedio', 'Num Transacciones', 'Unidades']
        df_cat = df_cat.reset_index()
        df_cat = df_cat.sort_values('Ingresos Totales', ascending=False)
        
        # Escribir datos
        ws['B2'] = "Ventas por Categoría"
        ws['B2'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ['Categoría', 'Ingresos Totales', 'Ticket Promedio', 'Transacciones', 'Unidades']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Datos
        for r_idx, row in enumerate(dataframe_to_rows(df_cat, index=False, header=False), start=5):
            for c_idx, value in enumerate(row, start=2):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Gráfico de barras
        chart = BarChart()
        chart.title = "Ingresos por Categoría"
        chart.x_axis.title = "Categoría"
        chart.y_axis.title = "Ingresos (€)"
        
        data = Reference(ws, min_col=3, min_row=4, max_row=4+len(df_cat))
        cats = Reference(ws, min_col=2, min_row=5, max_row=4+len(df_cat))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "B15")
        
        print("   ✅ Hoja Categorías creada")
    
    def _crear_hoja_tendencias(self, wb):
        """Crear hoja con tendencias temporales."""
        ws = wb.create_sheet("📈 Tendencias")
        
        # Preparar datos mensuales
        df_mes = self.df.groupby('mes').agg({
            'total': 'sum',
            'transaccion_id': 'count'
        }).reset_index()
        df_mes.columns = ['Mes', 'Ingresos', 'Transacciones']
        
        # Escribir datos
        ws['B2'] = "Tendencias Mensuales"
        ws['B2'].font = Font(size=14, bold=True)
        
        for r_idx, row in enumerate(dataframe_to_rows(df_mes, index=False, header=True), start=4):
            for c_idx, value in enumerate(row, start=2):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 4:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Gráfico de línea
        chart = LineChart()
        chart.title = "Evolución de Ingresos Mensuales"
        chart.x_axis.title = "Mes"
        chart.y_axis.title = "Ingresos (€)"
        
        data = Reference(ws, min_col=3, min_row=4, max_row=4+len(df_mes))
        cats = Reference(ws, min_col=2, min_row=5, max_row=4+len(df_mes))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "F4")
        
        print("   ✅ Hoja Tendencias creada")
    
    def _crear_hoja_top_productos(self, wb):
        """Crea hoja con top productos."""
        ws = wb.create_sheet("🏆 Top Productos")
        
        # Top 20 productos
        df_prod = self.df.groupby('producto_nombre')['total'].sum().nlargest(20).reset_index()
        df_prod.columns = ['Producto', 'Ingresos']
        df_prod['Ranking'] = range(1, len(df_prod) + 1)
        df_prod = df_prod[['Ranking', 'Producto', 'Ingresos']]
        
        # Título
        ws['B2'] = "Top 20 Productos Más Vendidos"
        ws['B2'].font = Font(size=14, bold=True)
        
        # Escribir datos con formato
        for r_idx, row in enumerate(dataframe_to_rows(df_prod, index=False, header=True), start=4):
            for c_idx, value in enumerate(row, start=2):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 4:  # Header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
                elif c_idx == 2 and r_idx <= 7:  # Top 3
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        
        # Ajustar anchos
        ws.column_dimensions['C'].width = 40
        
        print("   ✅ Hoja Top Productos creada")
    
    def _crear_hoja_datos(self, wb):
        """Crea hoja con muestra de datos crudos."""
        ws = wb.create_sheet("📄 Datos")
        
        # Muestra de 100 registros
        df_sample = self.df.head(100)[['fecha', 'categoria', 'producto_nombre', 
                                       'cantidad', 'precio_unitario', 'total', 'ciudad_envio']]
        
        # Escribir
        for r_idx, row in enumerate(dataframe_to_rows(df_sample, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=2):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 2:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        
        print("   ✅ Hoja Datos creada")


def main():
    """Función principal."""
    print("="*70)
    print("📊 GENERADOR AUTOMÁTICO DE REPORTES EXCEL")
    print("="*70)
    
    generator = ExcelReportGenerator()
    filepath = generator.crear_reporte_ejecutivo()
    
    print("\n" + "="*70)
    print("✨ REPORTE GENERADO EXITOSAMENTE")
    print("="*70)
    print(f"\n📁 Archivo: {filepath}")
    print("\n💡 Abre el archivo en Excel para ver el reporte completo")
    print("   - 5 hojas con diferentes análisis")
    print("   - KPIs formateados")
    print("   - Gráficos integrados")
    print("   - Formato profesional")


if __name__ == "__main__":
    main()
